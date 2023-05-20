"""
# Excel handler for test results reading and writing, settings for tests
# DK Jan 2023
"""


from typing import Tuple, List, Any
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import os
import time
import re
from datetime import datetime
from pprint import pprint
from dataclasses import dataclass

VERSION = "A.00.02"


@dataclass(frozen=True)
class DCV_Settings:
    function: str
    channel: str | int  # Allow for EXT
    coupling: str
    scale: float
    voltage: float
    offset: float
    bandwidth: int
    impedance: str
    invert: bool


@dataclass(frozen=True)
class Timebase_Settings:
    function: str
    channel: int
    timebase: float
    impedance: str | int
    bandwidth: int


@dataclass(frozen=True)
class Trigger_Settings:
    function: str
    channel: int
    scale: float
    voltage: float
    impedance: str | int
    frequency: float
    edge: str


@dataclass(frozen=True)
class Sampling_Settings:
    function: str
    channel: int
    coupling: str
    scale: float
    voltage: float
    timebase: float
    sample_rate: float
    frequency: float


@dataclass(frozen=True)
class Threshold_Settings:
    function: str
    channel_start: int
    voltage: float
    polarity: str


@dataclass
class Cell:
    col: int
    row: int
    value: Any


class ExcelInterface:
    """ """

    __filename: str = ""
    __start_row: int = 10
    __max_row: int = 1000
    __saved: bool = True
    __data_col = 10
    __result_col = 4
    __units_col = 6

    row: int = 1

    supported_test_names = [
        "BAL",
        "DCV",
        "DCV-BAL",
        "POS",
        "CURS",
        "RISE",
        "TIME",
        "TRIG",
        "IMP",
        "NOISE",
        "DELTAT",
    ]  # In order of test sequence preference - need list instead of set

    def __init__(self, filename, sheetindex=0) -> None:
        self.__filename = filename
        self.wb = openpyxl.load_workbook(
            self.__filename, read_only=False, data_only=False
        )
        self.ws = self.wb.worksheets[sheetindex]  # Default is the first sheet
        self.initialize()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.close()

    def read_result(self) -> float | Any:
        """
        read_result
        Special to read the data in the result column, for testing

        Returns:
            float: _description_
        """

        return self.ws.cell(column=self.__result_col, row=self.row).value

    def close(self) -> None:  # sourcery skip: raise-specific-error
        """
        close _summary_

        Raises:
            Exception: _description_
        """
        if not self.__saved:
            self.save_sheet()

        if not self.__saved:
            raise Exception("Unable to save")

        self.wb.close()

    def initialize(self) -> None:
        """
        initialize
        Set the row back to the start
        Multiple ways of initializing are in use:
        * Named Range StartCell which content is first row in data range
        * First row in the __data_col is the row
        * Default value in __start_row
        """

        if nr := self.get_named_cell("StartCell"):
            self.row = nr.row  # type: ignore
            self.__data_col = nr.col  # type: ignore
        elif rw := self.ws.cell(column=self.__data_col, row=1).value:
            try:
                self.row = int(rw)  # type: ignore
            except ValueError:
                self.row = self.__start_row
        else:
            self.row = self.__start_row

    def check_excel_available(self) -> bool:
        """
        check_excel_available
        Try to write to the file, if there is an exception user probably has it open
        """

        available = True

        try:
            self.wb.save(self.__filename)
        except PermissionError:
            available = False

        return available

    def check_valid_results(self) -> bool:
        """
        check_valid_results
        Check the model matches the expected values

        Returns:
            bool: _description_
        """

        # TODO make sure start cell defined
        nr = self.get_named_cell("StartCell")

        return bool(nr)

    def backup(self) -> None:
        """
        backup
        If the software crashes while the Excel instance is still open, it can corrupt the sheet
        """

        head, tail = os.path.split(self.__filename)

        backup_path = os.path.join(head, "Backups")
        fname = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{tail}"
        backup_name = os.path.join(backup_path, fname)

        if not os.path.exists(backup_path):
            os.mkdir(backup_path)

        self.wb.save(backup_name)

    def save_sheet(self) -> None:
        """
        save_sheet _summary_
        """
        try:
            # sometimes it throws an error if too quick
            self.wb.save(self.__filename)
            self.__saved = True
        except Exception:
            time.sleep(1)
            self.__saved = False

    def parse_value(self, val: str | float | int) -> str | float | int:
        """
        parse_value
        Some values have units appended, so convert to unit value

        Args:
            val (str): _description_

        Returns:
            float: _description_
        """

        if type(val) is not str:
            return val

        val = val.strip()
        if val[-1].isalpha():
            try:
                num = float(val[:-1])
            except ValueError:
                return val

            multiplier = val[-1]

            if multiplier == "G":
                num *= 1000000000
            elif multiplier == "M":
                num *= 1000000
            elif multiplier == "k":
                num *= 1000
            elif multiplier == "m":
                num /= 1000
            elif multiplier == "u":
                num /= 1000000
            elif multiplier == "n":
                num /= 1000000000
            elif multiplier == "p":
                num /= 1e12

            return num

        return val

    def get_column_row_number(self, coord: str) -> Tuple[int, int]:
        """
        get_column_number
        From the column letter(s), get the number

        Args:
            coord (str): "A1" to "XFD1048576"

        Returns:
            int: 1 based column number
        """

        xy = coordinate_from_string(coord)  # returns ('col', row)
        col: int = column_index_from_string(xy[0])
        row: int = xy[1]

        return [col, row]  # type: ignore

    def get_named_cell(self, name: str) -> Cell:
        """
        get_named_cell
        Get the cell coordinates from a named range.
        Although range can refer a range of cells, only the top left is returned
        as this is most commonly used in these applications

        Args:
            name (str): named range

        Returns:
            Tuple: col and row numbers
        """

        try:
            rng = self.wb.defined_names[name]  # type: ignore

            # This is returned in format tabname!cell_range

            # Create a generator
            dests = rng.destinations

            tabname, coord = next(dests)

            cd = self.get_column_row_number(coord)

            return Cell(col=cd[0], row=cd[1], value=self.ws[coord].value)  # type: ignore
        except KeyError:
            return None  # type: ignore

    def get_number_all_tests(self) -> int:
        """
        get_number_all_tests Get the total number of tests to perform, for the progress updates

        Returns:
            int: Number of rows with setup data
        """

        number_tests = 1  # Start row is pointing to first test

        self.initialize()

        while self.get_next_row():
            number_tests += 1

        self.initialize()  # Make sure row is reset

        return number_tests

    def get_next_row(self, supported_only: bool = True) -> bool:
        """
        Move down to the next row containing parameters

        Returns:
            bool: False when reached end of settings
        """
        valid = True

        self.row += 1

        while True:
            if val := self.ws.cell(column=self.__data_col, row=self.row).value:
                if str(val).upper() in self.supported_test_names and supported_only:
                    break

                if not supported_only:
                    break

            self.row += 1
            if self.row >= self.__max_row:
                valid = False
                break
        if self.row >= self.__max_row:
            valid = False

        return valid

    def get_test_rows(self, test_filter: str = "*") -> List:
        """
        get_test_rows
        Get a list of all test rows

        Returns:
            List: list of row numbers
        """

        self.initialize()

        test_rows = []

        while True:
            # match against the filter
            # Use a filter, but . is anything not *
            test_filter = test_filter.replace("*", ".")
            if test_filter.find(".") == -1:
                # Not using wildcard, make exact match on whole word
                test_filter = f"^{test_filter}$"
            setting = self.get_volt_settings()
            if re.match(test_filter, setting.function):  # type: ignore
                test_rows.append(self.row)

            if not self.get_next_row():
                break

        return test_rows

    def get_tb_test_settings(self, row: int = -1) -> Timebase_Settings:
        """
        get_tb_test_settings
        Get the settings relevant to the timebase test

        Args:
            row (int, optional): _description_. Defaults to -1.

        Returns:
            NamedTuple: _description_
        """

        if row == -1:
            row = self.row

        col = self.__data_col
        func = str(self.ws.cell(column=col, row=row).value)
        col += 1
        channel = self.ws.cell(column=col, row=row).value
        col += 1
        tb = self.ws.cell(column=col, row=row).value
        col += 1
        impedance = self.ws.cell(column=col, row=row).value
        col += 1
        bandwidth = self.ws.cell(column=col, row=row).value

        return Timebase_Settings(
            function=func,
            channel=channel,  # type: ignore
            timebase=tb,  # type: ignore
            impedance=impedance,  # type: ignore
            bandwidth=bandwidth,  # type: ignore
        )

    def get_trigger_settings(self, row: int = -1) -> Trigger_Settings:
        """
        get_trigger_settings
        Read the settings for the specified or current ro for trigger sensitivty test

        Args:
            row (int, optional): _description_. Defaults to -1.

        Returns:
            Trigger_Settings: _description_
        """

        if row == -1:
            row = self.row

        col = self.__data_col
        func = str(self.ws.cell(column=col, row=row).value)
        col += 1
        channel = self.ws.cell(column=col, row=row).value
        col += 1
        scale = self.ws.cell(column=col, row=row).value
        col += 1
        voltage = self.ws.cell(column=col, row=row).value
        col += 1
        impedance = self.ws.cell(column=col, row=row).value
        col += 1
        frequency = self.ws.cell(column=col, row=row).value
        col += 1
        edge = self.ws.cell(column=col, row=row).value

        edge_select = "F" if edge and edge.lower() == "f" else "R"  # type: ignore

        return Trigger_Settings(
            function=func,
            channel=channel,  # type: ignore
            scale=scale,  # type: ignore
            voltage=voltage,  # type: ignore
            impedance=impedance,  # type: ignore
            frequency=frequency,  # type: ignore
            edge=edge_select,
        )

    def get_volt_settings(self, row: int = -1) -> DCV_Settings:
        """
        get_volt_settings
        Read the current row as a test setting

        Args:
            row (_type_, optional): _description_. Defaults to self.row.

        Returns:
            NamedTuple: _description_
        """

        if row == -1:
            row = self.row

        col = self.__data_col
        func = str(self.ws.cell(column=col, row=row).value)
        col += 1
        chan = self.ws.cell(column=col, row=row).value
        col += 1
        coupling = self.ws.cell(column=col, row=row).value
        col += 1
        scale = self.ws.cell(column=col, row=row).value
        col += 1
        voltage = self.ws.cell(column=col, row=row).value
        col += 1
        offset = self.ws.cell(column=col, row=row).value
        col += 1
        bandwidth = self.ws.cell(column=col, row=row).value
        col += 1
        impedance = self.ws.cell(column=col, row=row).value
        col += 1
        invert = str(self.ws.cell(column=col, row=row).value)
        inverted = bool(invert and invert.lower() == "y") or invert == "1"

        return DCV_Settings(
            function=func,
            channel=chan,  # type: ignore
            coupling=coupling,  # type: ignore
            scale=scale,  # type: ignore
            voltage=voltage,  # type: ignore
            offset=offset,  # type: ignore
            bandwidth=bandwidth,  # type: ignore
            impedance=impedance,  # type: ignore
            invert=inverted,
        )

    def get_sample_rate_settings(self, row: int = -1) -> Sampling_Settings:
        """
        get_sample_rate_settings _summary_

        Args:
            row (int, optional): _description_. Defaults to -1.

        Returns:
            Sampling_Settings: _description_
        """

        if row == -1:
            row = self.row

        col = self.__data_col
        func = str(self.ws.cell(column=col, row=row).value)
        col += 1
        chan = self.ws.cell(column=col, row=row).value
        col += 1
        coupling = self.ws.cell(column=col, row=row).value
        col += 1
        scale = self.ws.cell(column=col, row=row).value
        col += 1
        voltage = self.ws.cell(column=col, row=row).value
        col += 1
        timebase = self.ws.cell(column=col, row=row).value
        col += 1
        sample_rate = self.parse_value(self.ws.cell(column=col, row=row).value)  # type: ignore
        col += 1
        frequency = self.parse_value(self.ws.cell(column=col, row=row).value)  # type: ignore

        return Sampling_Settings(
            function=func,
            channel=chan,  # type: ignore
            coupling=coupling,  # type: ignore
            scale=scale,  # type: ignore
            voltage=voltage,  # type: ignore
            timebase=timebase,
            sample_rate=sample_rate,
            frequency=frequency,
        )

    def get_all_test_settings(self, test_filter: str = "*") -> List:
        """
        get_all_test_settings
        Get a list of all test settings

        Returns:
            Dict: _description_
        """

        self.initialize()

        tests = []

        while True:
            setting = self.get_volt_settings()
            test_filter = test_filter.replace("*", ".")
            if test_filter.find(".") == -1:
                # Use exact match on whole word
                test_filter = f"^{test_filter}$"

            if re.match(test_filter, setting.function):  # type: ignore
                tests.append(setting)

            if not self.get_next_row():
                break

        return tests

    def get_test_types(self) -> set:
        """
        get_test_types
        Get a list of the unique tests

        Args:
            self (_type_): _description_
        """

        self.initialize()

        test_types = set()

        while True:
            if test_name := self.ws.cell(column=self.__data_col, row=self.row).value:
                if test_name not in []:  # TODO list of ignore tests if required
                    test_types.add(test_name)

            if not self.get_next_row():
                break

        return test_types

    def get_invalid_tests(self) -> List:
        """
        get_invalid_tests
        Return a list of rows where the test typoe is not one of the supported tests

        Returns:
            List: _description_
        """

        self.initialize()

        invalid_rows = []

        while True:
            if test_name := self.ws.cell(column=self.__data_col, row=self.row).value:
                if test_name not in self.supported_test_names:
                    invalid_rows.append([test_name, self.row])

            if not self.get_next_row(supported_only=False):
                break

        return invalid_rows

    def get_units(self) -> str:
        """
        get_units
        get the units

        Returns:
            str: _description_
        """

        return self.ws.cell(column=self.__units_col, row=self.row).value  # type: ignore

    def find_units_col(self, row: int = -1) -> int:
        """
        find_results_col
        Read the row one above current row and look for the column which has results or measured in it

        Args:
            row (int, optional): _description_. Defaults to -1.

        Returns:
            int: _description_
        """

        if row == -1:
            row = self.row

        row_count = 0

        while True:
            # some have comment rows, so keep going backwards until found
            row -= 1
            row_count += 1

            if row_count >= 5 or row < 10:  # All sheets have header rows
                return 0

            for col in range(1, 10):
                heading = str(self.ws.cell(column=col, row=row).value).lower()

                if "unit" in heading:
                    self.__units_col = col
                    return

            # not found

    def find_results_col(self, row: int = -1) -> int:
        """
        find_results_col
        Read the row one above current row and look for the column which has results or measured in it

        Args:
            row (int, optional): _description_. Defaults to -1.

        Returns:
            int: _description_
        """

        if row == -1:
            row = self.row

        row_count = 0

        while True:
            # some have comment rows, so keep going backwards until found
            row -= 1
            row_count += 1

            if row_count >= 5 or row < 10:  # All sheets have header rows
                return 0

            for col in range(1, 10):
                heading = str(self.ws.cell(column=col, row=row).value).lower()

                if "result" in heading or "measured" in heading:
                    return col

            # not found

    def write_result(
        self, result: float | str, save: bool = True, col: int = 0
    ) -> None:
        """
        write_result
        Write the data to the sheet at the current row

        Args:
            result (_type_): _description_
        """

        res_col = col or self.__result_col
        self.ws.cell(column=res_col, row=self.row).value = result

        if save:
            self.save_sheet()

    def write_data(self, data: float | int | str, named_range: str) -> bool:
        """
        write_data

        Write generic data to the named range

        If named range doesn't exist, nothing is written

        Args:
            data (float | int | str): _description_
            named_range (str): _description_

        Returns:
            bool: data written
        """

        if nr := self.get_named_cell(named_range):
            self.ws.cell(column=nr.col, row=nr.row).value = data
            self.save_sheet()
            return True

        return False


if __name__ == "__main__":
    with ExcelInterface("testsheets\\666_Tektronix_TDS3034C.xlsx") as excel:
        excel.backup()
        start_cell = excel.get_named_cell("StartCell")
        print(start_cell)
        # excel.row = start_cell.row
        # print(excel.get_named_cell("InvalidName"))
        print(f"Start row {excel.row}")
        print(f"Number tests {excel.get_number_all_tests()}")
        pprint(excel.get_test_rows())

        excel.initialize()

        pprint(excel.get_volt_settings())

        pprint(excel.get_all_test_settings())

        print("Filtered:")

        pprint(excel.get_all_test_settings("DC*"))

        print(f"Available: {excel.check_excel_available()}")

        pprint(excel.get_test_types())

        rows = excel.get_test_rows("TIME")
        if len(rows):
            pprint(rows)
            setting = excel.get_tb_test_settings(rows[0])
            print(setting.timebase)

        rows = excel.get_test_rows("TRIG")
        pprint(rows)
        if len(rows):
            settings = excel.get_trigger_settings(rows[0])
            pprint(settings)

        for name in excel.supported_test_names:
            print(name)

        print(excel.find_results_col(20))
        print(excel.find_results_col(68))
        print(excel.find_results_col(75))
        print(excel.find_results_col(80))

        print(excel.get_invalid_tests())
